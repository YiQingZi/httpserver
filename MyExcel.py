__author__ = 'guozixiang'
__version__ = "1.0.0"
# @createtime:2021/7/29 - 16:04


import os
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.styles import Alignment, Border ,Side ,colors
except ImportError as err:
    os.system('pip3 install openpyxl')
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.styles import Alignment, Border ,Side ,colors
from hadata import str_Y_M_D_H_M_s



fileName = ''


class Report:
    def __init__(self, name):
        global fileName
        self.name = f'{name}-{str_Y_M_D_H_M_s()}.xlsx'
        path = f'./result/'
        if not os.path.exists(path):
            os.mkdir(path)
        self.path = path + self.name
        fileName = self.path
        wb = Workbook()
        ws = wb.active
        ws.title = 'WeekReport'
        wb.save(self.path)
        wb.close()


    def ReportRead(self, sheet='WeekReport', cells='A', x=1):
        """读取表格基础方法"""
        try:
            wb = load_workbook(self.path)
            sh = wb[sheet]
            while True:
                result = sh[f'{cells}{x}'].value
                x += 1
                if None is not result:
                    yield result
                else:
                    wb.close()
                    break
        except IOError as err:
            return None


    def ReportResultWrite(self, sheet='WeekReport', cells='A2' , text='test'):
        """表格写入基础方法"""
        try:
            wb = load_workbook(self.path)
            sh = wb[sheet]
            try:
                sh[cells] = str(text)
            except ValueError:
                text1 = bytes(str(text),'gbk').decode('utf-8')
                sh[cells] = text1
            wb.save(self.path)
            wb.close()
            return True
        except (IOError,openpyxl.utils.exceptions.IllegalCharacterError) as err:
            return False


    def weekReport(self,sheet='WeekReport',len_project=1,task='',times='',remaks='',data=''):
        # ['slm321,5','slm333,6']
        abc = self._get_abc(len_project + 5)
        abc_1 = self._get_abc(len_project+ 5 + 1)
        try:
            wb = load_workbook(self.path)
            sh = wb[sheet]
            sh.merge_cells(f'E1:{abc}1')
            sh.merge_cells('A2:A3')
            sh.merge_cells('B2:B3')
            sh.merge_cells('C2:C3')
            sh.merge_cells('D2:D3')
            sh.merge_cells(f'{abc_1}2:{abc_1}3')
            sh.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
            sh.cell(2, 1).alignment = Alignment(horizontal='center', vertical='center')
            sh['A1'] = '日期'
            sh['B1'] = '本周计划'
            sh['C1'] = '本周实际工作情况'
            sh['D1'] = '本周问题'
            sh['E1'] = '人力统计'
            sh[f'{abc_1}1'] = '下周计划'

            sh['A2'] = times
            sh['C2'] = task
            sh[f'{abc_1}2'] = remaks
            for x in range(len_project):
                data_test = data[x].split(',')
                sh[f'{self._get_abc(5 + x)}2'] = data_test[0]
                sh[f'{self._get_abc(5 + x)}3'] = data_test[1]
            sh[f'{self._get_abc(len_project + 4 + 1)}2'] = '合计'

            border_set = Border(left=Side(style='medium', color=colors.BLACK),
                                right=Side(style='medium', color=colors.BLACK),
                                top=Side(style='medium', color=colors.BLACK),
                                bottom=Side(style='medium', color=colors.BLACK))
            for x in range(len_project+ 5 + 1):
                for y in range(3):
                    sh.cell(y+1, x+1).border = border_set

            wb.save(self.path)
            wb.close()
            return self.path
        except (IOError,openpyxl.utils.exceptions.IllegalCharacterError) as err:
            return None

    def _get_abc(self,nub):
        ABC_26 = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        if nub <= 26:
            return ABC_26[nub-1]
        if 52 >= nub > 26:
            return f'A{ABC_26[nub-26-1]}'
        if 78 >= nub > 52:
            return f'B{ABC_26[nub - 26 - 26 - 1]}'
        if 104 >= nub > 78:
            return f'C{ABC_26[nub - 26 - 26 -26- 1]}'
        if 130 >= nub > 104:
            return f'D{ABC_26[nub - 26 - 26 - 26 - 26 - 1]}'
        if 156 >= nub > 130:
            return f'D{ABC_26[nub - 26 - 26 - 26 - 26 - 26- 1]}'
        return None

def save_cvs(project,date,data):
    try:
        path = f"./result/{project}-{date}.csv"
        with open(path,"w+") as f:
            for x in data:
                #x = str(x).replace('<br>','\r\n')
                f.write(x)
                f.write('\n')
    except IOError:
        return None
    return path